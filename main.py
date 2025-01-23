from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import pandas as pd
import matplotlib.pyplot as plt


with open('data.txt', 'r') as file:
    data = json.load(file)


wb = Workbook()
ws = wb.active
ws.title = "Student Scores"

headers = ['Name'] + [item.title() for item in list(data['Joe'].keys())]
ws.append(headers)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)


headers = [cell.value for cell in ws[1]]
columns = [col for col, header in enumerate(headers, start=1) if header != 'Name' and header != 'Gender']

Percent = len(headers)+1 
ws.cell(row=1, column=Percent, value='Percent')

for row in range(2, ws.max_row + 1):
    scores = []
    
    for col in columns:
        score_value = ws.cell(row=row, column=col).value
        if score_value is not None: 
            scores.append(score_value)

    Precentage = sum(scores) / len(scores)

    ws.cell(row=row, column=Percent, value=Precentage)
    
wb.save("Student_scores.xlsx")

df = pd.read_excel('Student_scores.xlsx')
print(df)

subjects = ['Math', 'Science', 'English', 'Gym']
df = df.melt(id_vars='Name', value_vars=subjects, var_name='Subject', value_name='Scores')
pivot = df.groupby('Subject').agg(Score=('Scores', 'mean'))
pivot.to_excel('report.xlsx')


wb = load_workbook('report.xlsx')
ws = wb['Sheet1']

min_column=wb.active.min_column
max_column=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row


barchart = BarChart()
data = Reference(ws, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
category = Reference(ws, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(category)
barchart.title = 'Avg marks by Subject'
barchart.style = 5

ws.add_chart(barchart, "D9")
wb.save('report.xlsx')

print(pivot)
