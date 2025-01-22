from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import json


with open('data.txt', 'r') as file:
    data = json.load(file)


wb = Workbook()
ws = wb.active
ws.title = "Student Scores"

headers = ['Name'] + [item.title() for item in list(data['Joe'].keys())]
ws.append(headers)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)


headers = [cell.value for cell in ws[1]]
columns = [col for col, header in enumerate(headers, start=1) if header != 'Name']

Percent = len(headers)+1 
ws.cell(row=1, column=Percent, value='Percent')

for row in range(2, ws.max_row + 1):
    scores = []
    
    for col in columns:
        score_value = ws.cell(row=row, column=col).value
        if score_value is not None: 
            scores.append(score_value)

    Precentage = sum(scores) / len(scores)

    ws.cell(row=row, column=Percent, value=Precentage)
    
wb.save("Student_scores.xlsx")
